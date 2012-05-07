from openpyxl.reader.excel import load_workbook
from sqlclasses import *

class SpreadsheetReader(object):

    columns = {
            "Name": 1, 
            "Postcode": 6,
            "Telephone": 7,
            "InLondon": 8,
            "VenueType": 9,
            "MondayTime": 10,
            "MondayCharge": 11,
            "TuesdayTime": 12,
            "TuesdayCharge": 13,
            "WednesdayTime": 14,
            "WednesdayCharge": 15,
            "ThursdayTime": 16,
            "ThursdayCharge": 17,
            "FridayTime": 18,
            "FridayCharge": 19,
            "SaturdayTime": 20,
            "SaturdayCharge": 21,
            "SundayTime": 22,
            "SundayCharge": 23
    }

    def __init__(self, filename):
        wb = load_workbook(filename)
        self.sheet = wb.get_sheet_by_name('Venues')

    def get_rows(self):
        row = 2
        while self.sheet.cell(row=row, column=1).value != None:
            yield row
            row += 1

    def get_data(self, row, columnname):
        col = SpreadsheetReader.columns[columnname]
        return self.sheet.cell(row=row, column=col).value

def create_venue(reader, row):
    t = reader.get_data(row, "VenueType")
    if t == "P":
        tnum = 1
    elif t == "B":
        tnum = 2
    elif t == "C":
        tnum = 3
    else:
        tnum = 4

    v = Venue(
            name=reader.get_data(row, "Name"),
            postcode=reader.get_data(row, "Postcode"),
            phone=reader.get_data(row, "Telephone"),
            london=0,
            type=tnum,
            latitude=-1,
            longitude=-1)

    return v

def add_charges(venue, reader, row):
    for i, day in enumerate(["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]):
        charge = reader.get_data(row, day+"Charge")
        p = VenuePrice(price=str(charge), day=i+1, venue=venue)

def add_times(venue, reader, row):
    for i, day in enumerate(["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]):
        time = reader.get_data(row, day+"Time")
        p = VenueTime(time=str(time), day=i+1, venue=venue)


reader = SpreadsheetReader("venues.xlsx")
errored_rows = []

for row in reader.get_rows():
    print "processing row: " + str(row)
    postcode = reader.get_data(row, "Postcode")
    matching = Venue.select(Venue.q.postcode==postcode)
    found = (matching.count() == 1)
    venue = False

    if found:
        venue = list(matching)[0]
    elif matching.count() > 1:
        print "Found multiple matches for " + postcode + ", on row " + str(row)
        errored_rows.append(row)
    else:
        venue = create_venue(reader, row)

    if venue:
        add_charges(venue, reader, row)
        add_times(venue, reader, row)

print "MULTIPLE ROWS:"
print errored_rows
